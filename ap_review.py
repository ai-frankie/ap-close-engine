"""
AP CLOSE REVIEWER — independent QA check on a PREPARED workbook.
Recomputes truth from raw data (AP_Invoices, Payments, GL, PO, Receipts),
reads a human-prepared close workbook, and flags where the preparer's numbers
are wrong — and diagnoses WHY.

Run: python ap_review.py [path-to-prepared-workbook.xlsx]
If no path given, builds AP_Close_Prepared_SAMPLE.xlsx with planted mistakes.
"""
import csv, os, sys, datetime as dt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

D = os.path.dirname(os.path.abspath(__file__))
AP_CONTROL_ACCT = "2000"


def fnum(x):
    try: return float(str(x).replace(",", ""))
    except (ValueError, TypeError): return None


def load(n, data_dir=None):
    p = os.path.join(data_dir or D, n)
    return list(csv.DictReader(open(p, encoding="utf-8"))) if os.path.exists(p) else []


def compute_truth(data_dir=None):
    """Recompute AP truth from raw CSVs."""
    d = data_dir or D
    invs = load("AP_Invoices.csv", d)
    pays = load("Payments.csv", d)
    gl = load("GL.csv", d)
    pos = load("PO.csv", d)
    receipts = load("Receipts.csv", d)

    paid = defaultdict(float)
    for p in pays:
        paid[p.get("Invoice_Number")] += fnum(p.get("Amount")) or 0

    ap_sub = round(sum(max((fnum(inv.get("Amount")) or 0) - paid.get(inv.get("Invoice_Number"), 0), 0)
                       for inv in invs), 2)

    def _net_cr(r):
        return (fnum(r.get("Credit")) or 0) - (fnum(r.get("Debit")) or 0)
    gl_ap = round(sum(_net_cr(r) for r in gl
                       if str(r.get("Account_Number")).strip() == AP_CONTROL_ACCT), 2)

    # GRNI
    invoiced_pos = {inv.get("PO_Number") for inv in invs}
    recv_by_po = defaultdict(float)
    for r in receipts:
        recv_by_po[r.get("PO_Number")] += fnum(r.get("Qty_Received")) or 0
    po_price = {r["PO_Number"]: fnum(r["Unit_Price"]) or 0 for r in pos}
    grni = round(sum(recv_by_po[po] * po_price.get(po, 0)
                     for po in recv_by_po if po not in invoiced_pos), 2)

    # duplicate count
    from ap_duplicates import find_duplicates
    try:
        dup_count = len(find_duplicates(d))
    except (ValueError, FileNotFoundError):
        dup_count = 0

    return {
        "ap subledger": ap_sub,
        "gl ap control": gl_ap,
        "grni accrual": grni,
        "duplicate count": dup_count,
    }


def _build_sample(path, truth):
    """Build AP_Close_Prepared_SAMPLE.xlsx with planted errors."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Close"
    for r in [
        ["AP Close — Prepared", ""],
        ["AP subledger", truth["ap subledger"] * 2],  # double-count error
        ["GL AP control acct 2000", truth["gl ap control"]],
        ["Variance (GL - Subledger)", truth["gl ap control"] - truth["ap subledger"] * 2],
        ["GRNI accrual", 0],  # omission — preparer missed the GRNI
        ["Duplicate count", 0],  # preparer missed the duplicates
    ]:
        ws.append(r)
    wb.save(path)


def run_review(workbook_path=None, data_dir=None):
    """Review a prepared workbook against recomputed truth.
    Returns (rows, fails)."""
    truth = compute_truth(data_dir)

    path = workbook_path or os.path.join(D, "AP_Close_Prepared_SAMPLE.xlsx")
    if not os.path.exists(path):
        _build_sample(path, truth)

    wb = load_workbook(path, data_only=True)
    labelled = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            label = val = None
            for c in row:
                if c.value is None:
                    continue
                if label is None and isinstance(c.value, str):
                    label = c.value.strip().lower()
                elif val is None and isinstance(c.value, (int, float)):
                    val = float(c.value)
            if label is not None and val is not None:
                labelled.append((label, val))

    def reported(key):
        for lab, v in labelled:
            if key in lab:
                return v
        return None

    rows = []
    fails = 0
    for key, exp in truth.items():
        got = reported(key)
        if got is None:
            rows.append((key, "(missing)", f"{exp:,.2f}", "MISSING",
                         "Metric not found in workbook"))
            fails += 1
            continue
        ok = abs(got - exp) <= 0.5
        diag = ""
        if not ok:
            fails += 1
            if exp != 0 and abs(got) < 0.01:
                diag = "Omitted — preparer reported 0 for a non-zero figure"
            elif exp != 0 and abs(got / exp - 2) < 0.02:
                diag = "~2x expected — double-counted (SUM hit a total row?)"
            else:
                diag = f"Off by {got - exp:,.2f}"
        rows.append((key, f"{got:,.2f}", f"{exp:,.2f}", "OK" if ok else "FAIL", diag))

    return rows, fails


def write_review_report(rows, out_path=None):
    """Write AP_Close_Review.xlsx from run_review() rows."""
    out = Workbook()
    ws = out.active
    ws.title = "Review"
    ws.append(["Metric", "Reported", "Expected (recomputed)", "Result", "Diagnosis"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
    for r in rows:
        ws.append(list(r))
    for col, w in zip("ABCDE", [20, 18, 22, 10, 46]):
        ws.column_dimensions[col].width = w
    path = out_path or os.path.join(D, "AP_Close_Review.xlsx")
    out.save(path)
    return path


if __name__ == "__main__":
    wb_path = sys.argv[1] if len(sys.argv) > 1 else None
    if wb_path is None:
        sample = os.path.join(D, "AP_Close_Prepared_SAMPLE.xlsx")
        if not os.path.exists(sample):
            truth = compute_truth()
            _build_sample(sample, truth)
            print("(built sample-with-errors:", os.path.basename(sample), ")\n")
        wb_path = sample

    rows, fails = run_review(wb_path)
    out = write_review_report(rows)
    print("AP CLOSE REVIEW —", os.path.basename(wb_path))
    print("OUTPUT:", out, "\n")
    for m, g, e, res, diag in rows:
        print(f"  [{res}] {m}: reported {g} / expected {e}"
              + (f"  -> {diag}" if diag else ""))
    print(f"\n{fails} issue(s) found." if fails
          else "\nClean — all reported figures tie to the recomputed truth.")
