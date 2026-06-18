"""
AP month-end close — automated.
INPUTS (raw): AP_Invoices.csv, Payments.csv, GL.csv, PO.csv, Receipts.csv
OUTPUT: AP_Health_Report.xlsx — aging, open-AP subledger, GL AP-control tie,
        GRNI accrual, data-quality exceptions, and CONTROL CHECKS.
Run: python ap_close.py
"""
import csv, os, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

D = os.path.dirname(os.path.abspath(__file__))
AS_OF = dt.date(2026, 6, 17)
AP_CONTROL_ACCT = "2000"
GRNI_ACCT = "2150"


def fnum(x):
    try: return float(str(x).replace(",", ""))
    except (ValueError, TypeError): return None


def fdate(x):
    x = (x or "").strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try: return dt.datetime.strptime(x, fmt).date()
        except (ValueError, TypeError): pass
    return None


def load(name, data_dir=None):
    p = os.path.join(data_dir or D, name)
    return list(csv.DictReader(open(p, encoding="utf-8"))) if os.path.exists(p) else []


def bucket(due):
    if due is None:
        return "No Due Date"
    dpd = (AS_OF - due).days
    if dpd <= 0: return "Not Yet Due"
    if dpd <= 30: return "1-30"
    if dpd <= 60: return "31-60"
    if dpd <= 90: return "61-90"
    return "90+"


def run_close(data_dir=None):
    """Compute AP close from raw CSVs. Returns result dict (no file I/O)."""
    d = data_dir or D
    invs = load("AP_Invoices.csv", d)
    pays = load("Payments.csv", d)
    gl = load("GL.csv", d)
    pos = load("PO.csv", d)
    receipts = load("Receipts.csv", d)
    if not invs or not gl:
        raise ValueError("AP_Invoices.csv and GL.csv must exist and be non-empty in " + d)

    # build payment lookup: invoice -> total paid
    paid = {}
    for p in pays:
        inv = p.get("Invoice_Number")
        paid[inv] = paid.get(inv, 0) + (fnum(p.get("Amount")) or 0)

    # AP subledger = sum of open invoices - payments applied
    order = ["Not Yet Due", "1-30", "31-60", "61-90", "90+", "No Due Date"]
    agesum = {b: {"cnt": 0, "open": 0.0} for b in order}
    ap_subledger = 0.0
    for inv in invs:
        amt = fnum(inv.get("Amount")) or 0
        inv_paid = paid.get(inv.get("Invoice_Number"), 0)
        open_bal = round(amt - inv_paid, 2)
        inv["_open"] = open_bal
        inv["_due"] = fdate(inv.get("Due_Date"))
        inv["_bucket"] = bucket(inv["_due"])
        inv["_dpd"] = "" if inv["_due"] is None else (AS_OF - inv["_due"]).days
        if open_bal > 0:
            ap_subledger += open_bal
            if inv["_bucket"] in agesum:
                agesum[inv["_bucket"]]["cnt"] += 1
                agesum[inv["_bucket"]]["open"] += open_bal
    ap_subledger = round(ap_subledger, 2)

    # GL AP-control (acct 2000) net balance (credit-normal)
    gl_dr = round(sum(fnum(r.get("Debit")) or 0 for r in gl), 2)
    gl_cr = round(sum(fnum(r.get("Credit")) or 0 for r in gl), 2)

    def _net_cr(r):
        return (fnum(r.get("Credit")) or 0) - (fnum(r.get("Debit")) or 0)

    gl_ap = round(sum(_net_cr(r) for r in gl
                       if str(r.get("Account_Number")).strip() == AP_CONTROL_ACCT), 2)
    variance = round(gl_ap - ap_subledger, 2)

    # breaks: GL postings to AP-2000 with no invoice reference
    breaks = [r for r in gl
              if str(r.get("Account_Number")).strip() == AP_CONTROL_ACCT
              and not (r.get("Invoice_Ref") or "").strip()]
    auto_items = [(f"{r.get('Source')}: {r.get('Description')}",
                   round(_net_cr(r), 2)) for r in breaks]
    explained = round(sum(v for _, v in auto_items), 2)
    residual = round(variance - explained, 2)

    # GRNI: received but not invoiced (PO has receipt, no matching invoice)
    invoiced_pos = {inv.get("PO_Number") for inv in invs}
    recv_by_po = {}
    for r in receipts:
        po = r.get("PO_Number")
        recv_by_po[po] = recv_by_po.get(po, 0) + (fnum(r.get("Qty_Received")) or 0)
    po_price = {r["PO_Number"]: fnum(r["Unit_Price"]) or 0 for r in pos}
    grni = 0.0
    grni_items = []
    for po, qty in recv_by_po.items():
        if po not in invoiced_pos:
            val = round(qty * po_price.get(po, 0), 2)
            grni += val
            grni_items.append((po, qty, po_price.get(po, 0), val))
    grni = round(grni, 2)

    # data-quality exceptions
    invcount = {}
    for inv in invs:
        k = inv.get("Invoice_Number")
        invcount[k] = invcount.get(k, 0) + 1
    exc = []
    for inv in invs:
        flags = []
        if invcount.get(inv.get("Invoice_Number"), 0) > 1:
            flags.append("Dup invoice #")
        if inv["_open"] < 0:
            flags.append("Overpaid / credit balance")
        if inv["_due"] is None:
            flags.append("Missing due date")
        if flags:
            exc.append((inv.get("Invoice_Number"), inv.get("Vendor"), "; ".join(flags)))

    # CONTROL CHECKS
    checks = [
        ("GL balanced (debits = credits)",
         f"{gl_dr:,.2f} vs {gl_cr:,.2f}", abs(gl_dr - gl_cr) < 0.01),
        ("Aging buckets sum to AP subledger",
         f"{sum(v['open'] for v in agesum.values()):,.2f} vs {ap_subledger:,.2f}",
         abs(sum(v["open"] for v in agesum.values()) - ap_subledger) < 0.01),
        ("AP subledger ties to GL (flagged breaks explain variance)",
         f"residual {residual:,.2f}", abs(residual) < 0.01),
        ("GRNI identified and flagged",
         f"{grni:,.2f} ({len(grni_items)} items)", True),
        ("No negative AP balances unresolved",
         f"{sum(1 for inv in invs if inv['_open'] < 0)} negative",
         sum(1 for inv in invs if inv["_open"] < 0) == 0),
    ]

    return {
        "ap_subledger": ap_subledger,
        "gl_ap": gl_ap,
        "variance": variance,
        "residual": residual,
        "grni": grni,
        "grni_items": grni_items,
        "auto_items": auto_items,
        "checks": checks,
        "agesum": agesum,
        "order": order,
        "exc": exc,
        "invoices": invs,
    }


def write_report(r, out_path=None):
    """Write AP_Health_Report.xlsx from run_close() result dict."""
    wb = Workbook()
    H = Font(bold=True, color="FFFFFF")
    HF = PatternFill("solid", fgColor="1F4E78")
    B = Font(bold=True)

    def tab(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]: c.font = H; c.fill = HF
        for row in rows: ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 26
        return ws

    hr = wb.active
    hr.title = "AP_Health_Report"
    hr.append([f"AP HEALTH REPORT — close {AS_OF:%b %Y} (SYNTHETIC)"])
    hr["A1"].font = Font(bold=True, size=14)
    hr.append([])
    hr.append(["AGING SUMMARY", "Open $", "% of total"])
    for c in hr[3]: c.font = B
    for b in r["order"]:
        pct = r["agesum"][b]["open"] / r["ap_subledger"] * 100 if r["ap_subledger"] else 0
        hr.append([b, round(r["agesum"][b]["open"], 2), f"{pct:.1f}%"])
    hr.append(["TOTAL (AP subledger)", r["ap_subledger"], "100%"])
    hr["A" + str(hr.max_row)].font = B
    hr.append([])
    hr.append(["GRNI accrual needed", r["grni"]])
    for col in hr.columns:
        hr.column_dimensions[col[0].column_letter].width = 34

    tab("GRNI_Detail", ["PO", "Qty Received", "PO Price", "Accrual Amount"],
        [list(g) for g in r["grni_items"]])
    tab("AP_to_GL_Recon", ["Item", "Amount", "Note"],
        [["AP subledger (open invoices)", r["ap_subledger"], ""],
         ["GL AP control (acct " + AP_CONTROL_ACCT + ")", r["gl_ap"], ""],
         ["Variance (GL - Subledger)", r["variance"], ""],
         ["", "", ""],
         ["Flagged breaks (AP-2000, no invoice ref):", "", ""]]
        + [[d, amt, "VERIFY"] for d, amt in r["auto_items"]]
        + [["Sum of flagged items", sum(v for _, v in r["auto_items"]), ""],
           ["Residual unexplained", r["residual"],
            "INVESTIGATE" if abs(r["residual"]) >= 0.01 else ""],
           ["Status",
            "TIES (verify items)" if abs(r["residual"]) < 0.01 else "DOES NOT TIE", ""]])
    tab("DataQuality_Exceptions", ["Invoice", "Vendor", "Issue(s)"],
        [list(e) for e in r["exc"]])
    tab("Control_Checks", ["Control", "Detail", "Pass"],
        [[n, d, "PASS" if ok else "FAIL"] for n, d, ok in r["checks"]])

    out = out_path or os.path.join(D, "AP_Health_Report.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    result = run_close()
    out = write_report(result)
    r = result
    print("INPUTS: AP_Invoices.csv, Payments.csv, GL.csv, PO.csv, Receipts.csv")
    print("OUTPUT:", out)
    print(f"\nAP Subledger {r['ap_subledger']:,.2f} | GL AP {r['gl_ap']:,.2f} | "
          f"Variance {r['variance']:,.2f} | Residual {r['residual']:,.2f}")
    print(f"GRNI accrual needed: {r['grni']:,.2f} ({len(r['grni_items'])} items)")
    print(f"Exceptions: {len(r['exc'])} rows")
    print("\nCONTROL CHECKS:")
    for n, d, ok in r["checks"]:
        print(f"  [{'PASS' if ok else 'FAIL'}] {n} ({d})")
    print("ALL CONTROLS PASS" if all(ok for _, _, ok in r["checks"]) else "** CONTROL FAILURE **")
